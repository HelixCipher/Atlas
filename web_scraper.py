import sys
import asyncio
import time
import logging
import os
import re
import sqlite3
import tempfile
from pathlib import Path
from bs4 import BeautifulSoup, Tag, NavigableString
from tqdm import tqdm
from playwright.async_api import async_playwright
from openpyxl import Workbook, load_workbook

# Configure logging with timestamps.
logging.basicConfig(level=logging.DEBUG,
                    format='%(asctime)s - %(levelname)s - %(message)s')

CONTENT_TIMEOUT = 30000   # in milliseconds
RETRY_COUNT = 3
RETRY_DELAY = 1           # in seconds
DEBUG_HTML_DIR = "debug_html"
os.makedirs(DEBUG_HTML_DIR, exist_ok=True)

# Regex pattern to match dates in the format "26 februari 2025"
date_pattern = re.compile(r'(\d{1,2}\s+\w+\s+\d{4})')

async def add_cookie_removal_script(context):
    await context.add_init_script(script="""
        const removeCookieBanner = () => {
            const selectors = ['div', 'p', 'span'];
            selectors.forEach(selector => {
                document.querySelectorAll(selector).forEach(el => {
                    if (el.innerText && el.innerText.includes('Vi använder kakor')) {
                        el.remove();
                    }
                });
            });
        };
        if (document.readyState === 'loading') {
            document.addEventListener('DOMContentLoaded', removeCookieBanner);
        } else {
            removeCookieBanner();
        }
    """)

async def get_html(page, url):
    try:
        await page.goto(url, timeout=CONTENT_TIMEOUT)
        await page.wait_for_load_state("networkidle", timeout=CONTENT_TIMEOUT)
        html = await page.content()
        return html
    except Exception as e:
        raise e

async def get_report_html_alternative(browser, url, retries=RETRY_COUNT):
    for attempt in range(1, retries + 1):
        page = None
        try:
            logging.debug(f"Attempt {attempt} to get content for {url}")
            page = await browser.new_page()
            html = await get_html(page, url)
            if html and len(html) > 100:
                logging.debug(f"Successfully retrieved content (length: {len(html)}) on attempt {attempt}")
                return html
            else:
                logging.error(f"Empty or too short content for {url} on attempt {attempt}")
        except Exception as e:
            logging.error(f"Error on attempt {attempt} for {url}: {e}")
        finally:
            if page is not None:
                try:
                    await page.close()
                except Exception as close_e:
                    logging.error(f"Error closing page for {url}: {close_e}")
        await asyncio.sleep(RETRY_DELAY)
    logging.error(f"All {retries} attempts failed for {url}.")
    return ""

def is_valid_listing_page(html):
    soup = BeautifulSoup(html, "html.parser")
    header = soup.find(lambda tag: tag.name in ['h1', 'h2', 'h3'] and "Publikationer" in tag.get_text())
    return header is not None

def extract_report_links(listing_html):
    """
    Extracts report URLs and their associated dates.
    Iterates over all <time> elements with class "lp-filterable-list-item-date" and:
      - Retrieves the date text.
      - Uses find_previous to locate the closest <a> element with href starting with '/publikationer/'.
    Returns a list of dictionaries with 'url' and 'date' keys.
    """
    soup = BeautifulSoup(listing_html, "html.parser")
    reports = []
    time_tags = soup.select("time.lp-filterable-list-item-date")
    logging.debug(f"Found {len(time_tags)} time tags on the page.")
    for time_tag in time_tags:
        date_str = time_tag.get_text(strip=True)
        # Find the closest preceding <a> element with href starting with "/publikationer/"
        link_tag = time_tag.find_previous("a", href=re.compile(r"^/publikationer/"))
        if not link_tag:
            logging.debug(f"No anchor found for time tag: {time_tag}")
            continue
        href = link_tag.get("href")
        if href:
            parts = href.split('/')
            if len(parts) > 2:
                category = parts[2].lower()
                if category in ["rapport", "pm", "statistik", "wp"]:
                    full_url = href if href.startswith("http") else "https://www.tillvaxtanalys.se" + href
                    logging.debug(f"Extracted report: {full_url} with date: {date_str}")
                    if not any(report["url"] == full_url for report in reports):
                        reports.append({"url": full_url, "date": date_str})
    logging.debug(f"Extracted {len(reports)} report links with dates: {reports}")
    return reports


def _find_label_value(soup: BeautifulSoup, label_base: str) -> str:
    """
    Locate the text corresponding to a label like "Serienummer" or "Diarienummer",
    handling cases where the HTML may have:
      - 'Serienummer' on one line, then ':' on the next line, then the real value.
      - 'Serienummer:' and the value on the same stripped string.
      - 'Serienummer:' alone on one stripped string, then ':' alone, then the value.
    Approach:
      1) Build a list of all stripped_strings.
      2) For each index i where stripped[i] matches one of:
         a) exactly label_base (e.g. "Serienummer")
         b) starts with label_base + ":" (e.g. "Serienummer:")
         do:
           - If the same stripped string already contains text after "label_base:", return it.
           - Otherwise, walk forward from i+1, skipping any stripped[j] that is exactly ":" (or empty),
             and return the first stripped[j] that is neither ":" nor blank.
      3) If no label is found at all, return None.
    """
    lb_lower = label_base.lower()
    stripped = list(soup.stripped_strings)
    n = len(stripped)

    for i, s in enumerate(stripped):
        sl = s.strip()
        slower = sl.lower()

        # (1) Case A: exact label without colon (e.g. "Serienummer")
        if slower == lb_lower:
            # Advance j until we find something other than just ":"
            j = i + 1
            while j < n and stripped[j].strip() == ":":
                j += 1
            if j < n:
                return stripped[j].strip()
            else:
                return None

        # (2) Case B: label with colon on same line (e.g. "Serienummer: Rapport 2024:xx")
        if slower.startswith(lb_lower + ":"):
            after = sl[len(label_base) + 1 :].strip()  # text after "Serienummer:"
            if after and after != ":":
                return after
            # If the stripped string is exactly "Serienummer:" or ends with "Serienummer: :",
            # we still need to move forward skipping any standalone ":".
            j = i + 1
            while j < n and stripped[j].strip() == ":":
                j += 1
            if j < n:
                return stripped[j].strip()
            else:
                return None

    # (3) If we reach here, neither pattern was found
    return None



def parse_report(report_html: str) -> dict:
    """
    Parses a report page and returns:
      - report_name (the <h1> title)
      - serienummer (e.g. "Rapport 2024:17")
      - diarienummer (e.g. "2021/68")
      - description (combined text from .rapport-description or all <p> tags)
    """
    soup = BeautifulSoup(report_html, "html.parser")

    # 1) Report name (the page’s <h1>)
    h1 = soup.find("h1")
    report_name = h1.get_text(strip=True) if h1 else None

    # 2) Serienummer & Diarienummer
    serienummer = _find_label_value(soup, "Serienummer")
    diarienummer = _find_label_value(soup, "Diarienummer")

    # 3) Description
    desc_container = soup.find("div", class_="rapport-description")
    if desc_container:
        description = desc_container.get_text(separator=" ", strip=True)
    else:
        article_container = soup.find("div", class_="rapport-article-content")
        if article_container:
            paragraphs = article_container.find_all("p")
        else:
            paragraphs = soup.find_all("p")
        description = " ".join([p.get_text(strip=True) for p in paragraphs])

    return {
        "report_name": report_name,
        "serienummer": serienummer,
        "diarienummer": diarienummer,
        "description": description,
    }





def save_to_excel(reports_data, filename="reports.xlsx"):
    if os.path.exists(filename):
        wb = load_workbook(filename)
        ws = wb.active
        existing_urls = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[-1]:
                existing_urls.add(row[-1])
        for report in reports_data:
            if report.get("url") not in existing_urls:
                row = [
                    report.get("report_name"),
                    report.get("diarienummer"),
                    report.get("serienummer"),
                    report.get("description"),
                    report.get("date"),
                    report.get("url"),
                ]
                ws.append(row)
        wb.save(filename)
        logging.info(f"Updated existing Excel file: {filename}")
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Reports"
        headers = ["report_name", "diarienummer", "serienummer", "description", "date", "url"]
        ws.append(headers)
        for report in reports_data:
            row = [
                report.get("report_name"),
                report.get("diarienummer"),
                report.get("serienummer"),
                report.get("description"),
                report.get("date"),
                report.get("url"),
            ]
            ws.append(row)
        wb.save(filename)
        logging.info(f"Created new Excel file: {filename}")

def save_to_sqlite(reports_data, db_filename="reports.db"):
    conn = sqlite3.connect(db_filename)
    cursor = conn.cursor()
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS reports (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            report_name TEXT,
            diarienummer TEXT,
            serienummer TEXT,
            description TEXT,
            date TEXT,
            url TEXT UNIQUE
        )
    """)
    for report in reports_data:
        try:
            cursor.execute("""
                INSERT INTO reports (report_name, diarienummer, serienummer, description, date, url)
                VALUES (?, ?, ?, ?, ?, ?)
            """, (
                report.get("report_name"),
                report.get("diarienummer"),
                report.get("serienummer"),
                report.get("description"),
                report.get("date"),
                report.get("url"),
            ))
        except sqlite3.IntegrityError:
            logging.info(f"Skipping duplicate entry for URL: {report.get('url')}")
    conn.commit()
    conn.close()
    logging.info(f"Saved data to SQLite database: {db_filename}")

async def fetch_listing_pages(listing_url):
    all_listing_htmls = []
    async with async_playwright() as p:
        with tempfile.TemporaryDirectory() as temp_dir:
            context = await p.chromium.launch_persistent_context(
                temp_dir,
                headless=True,
                args=["--headless=new"],
            )
            await context.add_cookies([{
                "name": "CONSENT",
                "value": "YES+",
                "domain": ".tillvaxtanalys.se",
                "path": "/"
            }])
            await add_cookie_removal_script(context)
            page = await context.new_page()
            try:
                current_page = 1
                while True:
                    if current_page == 1:
                        paginated_url = listing_url
                    else:
                        paginated_url = listing_url + f"&svAjaxReqParam=ajax&page12_706c70df1932999ea346c0a={current_page}"
                    logging.info(f"Fetching listing page: {paginated_url}")
                    html = await get_html(page, paginated_url)
                    if not is_valid_listing_page(html):
                        logging.warning(f"'Publikationer' not found on {paginated_url}. Assuming last page reached.")
                        break
                    if not html or len(html) < 1000:
                        logging.warning(f"Received empty or too short content for {paginated_url}. Assuming no more pages.")
                        break
                    all_listing_htmls.append(html)
                    report_links = extract_report_links(html)
                    if len(report_links) < 1:
                        logging.info("Less than expected reports on this page, assuming last page.")
                        break
                    current_page += 1
                    await asyncio.sleep(0.5)
            finally:
                await page.close()
                await context.close()
    return all_listing_htmls

async def main():
    listing_url = ("https://www.tillvaxtanalys.se/publikationer.109.html?"
                   "sv.target=12.706c70df1932999ea346c0a&"
                   "sv.12c70df1932999ea346c0a.route=/&"
                   "query=&from=2000-01-01&to=2025-01-01")
    logging.info(f"Fetching listing pages from: {listing_url}")
    listing_htmls = await fetch_listing_pages(listing_url)
    
    report_entries = []
    for html in listing_htmls:
        report_entries.extend(extract_report_links(html))
    logging.info(f"Found {len(report_entries)} report entries: {report_entries}")
    
    async with async_playwright() as p:
        with tempfile.TemporaryDirectory() as temp_dir:
            context = await p.chromium.launch_persistent_context(
                temp_dir,
                headless=True,
                args=["--headless=new"],
            )
            await context.add_cookies([{
                "name": "CONSENT",
                "value": "YES+",
                "domain": ".tillvaxtanalys.se",
                "path": "/"
            }])
            await add_cookie_removal_script(context)
            
            reports_data = []
            for entry in tqdm(report_entries, desc="Processing reports"):
                url = entry.get("url")
                list_date = entry.get("date")
                logging.debug(f"Processing report with URL: {url} and listing date: {list_date}")
                try:
                    page = await context.new_page()
                    html = await get_html(page, url)
                    if len(html) < 1000:
                        debug_file = os.path.join(DEBUG_HTML_DIR, url.split("/")[-1] + ".html")
                        with open(debug_file, "w", encoding="utf-8") as f:
                            f.write(html)
                        logging.warning(f"Content for {url} is very short. Saved HTML to {debug_file}")
                    data = parse_report(html)
                    data["url"] = url
                    data["date"] = list_date  # Use the date from the listing
                    reports_data.append(data)
                    logging.debug(f"Completed processing report: {data}")
                    await page.close()
                except Exception as e:
                    logging.error(f"Error processing {url}: {e}")
                await asyncio.sleep(0.5)
            await context.close()
    
    logging.info("Final Report Data:")
    for report in reports_data:
        logging.info(report)
    
    save_to_excel(reports_data)
    save_to_sqlite(reports_data)

if __name__ == "__main__":
    asyncio.run(main())
